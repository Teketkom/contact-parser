"""
Модуль экспорта результатов парсинга в многолистовой Excel (XLSX).

Структура файла:
  - По одному листу на каждую CRM-категорию (только непустые).
  - Лист «Другие категории» для некатегоризированных контактов.
  - Лист «Статистика» — сводная информация.
  - Лист «Лог» — лог обработки по сайтам.

Соответствует требованиям ТЗ:
  - FR-OUT-001: набор столбцов
  - FR-OUT-002: автофильтр, закреплённый заголовок
  - FR-OUT-004: лист Статистика
"""

from __future__ import annotations

import logging
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.core.categorizer import categorize_contacts
from app.models import ContactRecord, ParseTask

logger = logging.getLogger(__name__)

# ── Цвета ─────────────────────────────────────────────────────────────────────

COLOR_HEADER_BG   = "1E3A5F"   # Тёмно-синий заголовок
COLOR_HEADER_FONT = "FFFFFF"   # Белый текст
COLOR_ROW_ODD     = "FFFFFF"   # Нечётные строки — белые
COLOR_ROW_EVEN    = "F0F4FA"   # Чётные строки — светло-голубые
COLOR_SUMMARY_BG  = "E8F5E9"   # Светло-зелёный для сводки
COLOR_STATS_ROW   = "EEF2F7"   # Светло-серый для меток в статистике

# ── Шрифты ────────────────────────────────────────────────────────────────────

FONT_HEADER  = Font(name="Calibri", size=10, bold=True, color=COLOR_HEADER_FONT)
FONT_BODY    = Font(name="Calibri", size=10)
FONT_LINK    = Font(name="Calibri", size=10, color="0563C1", underline="single")
FONT_SUMMARY = Font(name="Calibri", size=11, bold=True)

# ── Определение столбцов листа контактов ──────────────────────────────────────

COLUMNS: list[dict[str, Any]] = [
    {"key": "company_name",        "header": "Название компании",              "width": 30},
    {"key": "site_url",            "header": "Сайт",                           "width": 35},
    {"key": "company_email",       "header": "Общий email",                    "width": 30},
    {"key": "position_raw",        "header": "Должность (как на сайте)",       "width": 35},
    {"key": "position_normalized", "header": "Должность (нормализованная)",    "width": 30},
    {"key": "full_name",           "header": "ФИО",                            "width": 28},
    {"key": "fio_dative",          "header": "Фамилия И.О. в дат. падеже",    "width": 25},
    {"key": "name_patronymic",     "header": "Имя Отчество",                   "width": 25},
    {"key": "gender_ending",       "header": "Окончание (ый/ой дл.)",          "width": 16},
    {"key": "personal_email",      "header": "Личный email",                   "width": 30},
    {"key": "phone",               "header": "Телефон",                        "width": 18},
    {"key": "inn",                 "header": "ИНН",                            "width": 14},
    {"key": "kpp",                 "header": "КПП",                            "width": 12},
    {"key": "source_url",          "header": "URL страницы-источника",         "width": 40},
    {"key": "page_language",       "header": "Язык страницы",                  "width": 12},
    {"key": "scan_date",           "header": "Дата сканирования",              "width": 20},
]


# ── Вспомогательные функции для ФИО ──────────────────────────────────────────

def _split_fio(full_name: str) -> tuple[str, str, str]:
    """
    Разбивает ФИО на (Фамилия, Имя, Отчество).

    Поддерживает:
    - Полный формат: «Калашников Дмитрий Владимирович»
    - Инициалы: «Калашников Д.В.» или «Калашников Д. В.»
    - Смешанный: «Калашников Дмитрий В.»
    """
    if not full_name or not full_name.strip():
        return ("", "", "")

    # Нормализуем пробелы
    text = re.sub(r"\s+", " ", full_name.strip())

    # Разбиваем токены: либо слово с заглавной, либо инициал вида Д.
    tokens = re.findall(r"[А-ЯЁA-Z][а-яёa-z]*\.?|[А-ЯЁ]\.", text)

    if not tokens:
        return (full_name.strip(), "", "")

    family = tokens[0] if len(tokens) >= 1 else ""
    name   = tokens[1] if len(tokens) >= 2 else ""
    patron = tokens[2] if len(tokens) >= 3 else ""

    return (family, name, patron)


def _detect_gender(full_name: str) -> str:
    """
    Определяет пол по окончанию отчества.

    - Оканчивается на «-вич» / «-ич» → мужской («м»)
    - Оканчивается на «-вна» / «-чна» → женский («ж»)
    - Не определено → «м» (по умолчанию)
    """
    if not full_name:
        return "м"

    _, _, patron = _split_fio(full_name)
    if not patron:
        # Попробуем определить по фамилии: Петрова/Иванова → «ж», Петров/Иванов → «м»
        family, _, _ = _split_fio(full_name)
        if family and family.endswith(("ова", "ева", "ина", "ая", "яя")):
            return "ж"
        return "м"

    patron_lower = patron.lower().rstrip(".")
    if patron_lower.endswith(("вич", "ич")):
        return "м"
    if patron_lower.endswith(("вна", "чна", "на")):
        return "ж"

    return "м"


def _gender_ending(gender: str) -> str:
    """
    Возвращает «ый» для мужского рода, «ой» для женского.
    Используется для «Уважаемый» / «Уважаемой».
    """
    return "ый" if gender == "м" else "ой"


def _to_dative_family(family: str, gender: str) -> str:
    """
    Склоняет фамилию в дательный падеж по базовым правилам русского языка.

    Правила:
    - Муж., оканч. на -ий / -ый → -ому (Достоевский → Достоевскому)
    - Муж., оканч. на -ой → -ому (Толстой → Толстому)
    - Муж., оканч. на согласный → +у (Калашников → Калашникову)
    - Жен., оканч. на -ая / -яя → -ой (Достоевская → Достоевской)
    - Жен., оканч. на -а → -е (Петрова → Петровой — особый случай на -ва/-ва)
    - Жен., оканч. на -ова / -ева / -ина → -ой (Петрова → Петровой)
    - Не изменяемые фамилии (на -о, -е, -и, -у, -ю) остаются.
    """
    if not family:
        return family

    f = family.rstrip(".")
    fl = f.lower()

    # Мужские
    if gender == "м":
        if fl.endswith("ий") or fl.endswith("ый"):
            return f[:-2] + "ому"
        if fl.endswith("ой"):
            return f[:-2] + "ому"
        if fl.endswith("ь"):
            return f[:-1] + "ю"
        # Оканчивается на согласную (типичный случай: Калашников, Смирнов)
        consonants = "бвгджзйклмнпрстфхцчшщ"
        if fl[-1] in consonants:
            return f + "у"
        # Неизменяемые (на -о, -е, -и, -у, -ю, -а без -ова/-ева/-ина)
        return f

    # Женские
    else:
        if fl.endswith("ая") or fl.endswith("яя"):
            return f[:-2] + "ой"
        if fl.endswith("ова") or fl.endswith("ева") or fl.endswith("ина"):
            return f[:-1] + "ой"
        if fl.endswith("а"):
            # Общий случай: убираем -а, добавляем -е (Нина → Нине, редкий для фамилий)
            return f[:-1] + "е"
        # Неизменяемые
        return f


def _to_dative(family: str, name_initial: str, patronymic_initial: str) -> str:
    """
    Переводит ФИО в дательный падеж формата «Фамилии И.О.».

    Пример: Калашников, Д, В → «Калашникову Д.В.»
    Инициалы остаются без изменений.
    """
    if not family:
        return ""

    # Определяем пол по отчеству-инициалу или по самой фамилии
    # Отчество в этой функции — уже инициал, пол надо передать извне;
    # делаем эвристику по фамилии
    fl = family.lower()
    if fl.endswith(("ова", "ева", "ина", "ая", "яя", "цкая", "ская")):
        gender = "ж"
    else:
        gender = "м"

    dative_family = _to_dative_family(family, gender)

    # Форматируем инициалы
    parts = []
    if name_initial:
        initial = name_initial.strip(".")[0].upper() + "." if name_initial.strip(".") else ""
        parts.append(initial)
    if patronymic_initial:
        initial = patronymic_initial.strip(".")[0].upper() + "." if patronymic_initial.strip(".") else ""
        parts.append(initial)

    initials_str = "".join(parts)

    if initials_str:
        return f"{dative_family} {initials_str}"
    return dative_family


def _extract_name_patronymic(full_name: str) -> str:
    """
    Извлекает «Имя Отчество» из полного ФИО.

    Пример: «Калашников Дмитрий Владимирович» → «Дмитрий Владимирович»
    Если отчество отсутствует, возвращает только имя.
    """
    _, name, patron = _split_fio(full_name)
    parts = [p for p in (name, patron) if p]
    return " ".join(parts)


# ── Служебные функции форматирования листа ────────────────────────────────────

def _cell_border() -> Border:
    """Тонкая граница ячеек."""
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _apply_header_row(ws, headers: list[str]) -> None:
    """Добавляет строку заголовка с форматированием."""
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    for cell in ws[1]:
        cell.font = FONT_HEADER
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _cell_border()
    ws.row_dimensions[1].height = 32


def _apply_column_widths(ws, columns: list[dict[str, Any]]) -> None:
    """Устанавливает ширину столбцов."""
    for i, col_def in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(i)].width = col_def["width"]


def _apply_autofilter_freeze(ws, last_col: int, last_row: int) -> None:
    """Включает автофильтр и закрепляет заголовок."""
    ws.freeze_panes = "A2"
    if last_row >= 2:
        ws.auto_filter.ref = f"A1:{get_column_letter(last_col)}{last_row}"


# ── Преобразование контакта в строку ─────────────────────────────────────────

def _contact_to_row(contact: ContactRecord) -> list[Any]:
    """Преобразует ContactRecord в список значений для строки листа."""

    # --- ФИО-производные ---
    full_name = contact.full_name or ""
    family, name, patron = _split_fio(full_name)
    gender = _detect_gender(full_name)

    # Инициалы для дательного падежа
    name_init   = name[0]   if name   else ""
    patron_init = patron[0] if patron else ""
    fio_dative  = _to_dative(family, name_init, patron_init)

    name_patronymic = _extract_name_patronymic(full_name)
    ending = _gender_ending(gender)

    # --- Дата сканирования ---
    scan_date_str = ""
    if contact.scan_date:
        try:
            scan_date_str = contact.scan_date.strftime("%d.%m.%Y %H:%M")
        except Exception:
            scan_date_str = str(contact.scan_date)

    # --- Нормализованная должность: первая буква заглавная ---
    pos_normalized = contact.position_normalized or ""
    if pos_normalized:
        pos_normalized = pos_normalized[0].upper() + pos_normalized[1:]

    return [
        contact.company_name or "",
        contact.site_url or "",
        contact.company_email or "",
        contact.position_raw or "",
        pos_normalized,
        full_name,
        fio_dative,
        name_patronymic,
        ending,
        contact.personal_email or "",
        contact.phone or contact.phone_raw or "",
        contact.inn or "",
        contact.kpp or "",
        contact.source_url or "",
        contact.page_language or "",
        scan_date_str,
    ]


# ── Построение листов ─────────────────────────────────────────────────────────

def _build_category_sheet(ws, contacts: list[ContactRecord]) -> None:
    """Заполняет один лист категории: заголовок + строки контактов."""
    headers = [col["header"] for col in COLUMNS]
    _apply_header_row(ws, headers)
    _apply_column_widths(ws, COLUMNS)

    # Ключи столбцов с гиперссылками
    link_keys = {"site_url", "source_url"}
    mailto_keys = {"company_email", "personal_email"}

    for row_idx, contact in enumerate(contacts, start=2):
        row_data = _contact_to_row(contact)
        ws.append(row_data)

        bg_color = COLOR_ROW_EVEN if row_idx % 2 == 0 else COLOR_ROW_ODD
        row_fill = PatternFill("solid", fgColor=bg_color)

        for col_idx, cell in enumerate(ws[row_idx], start=1):
            col_key = COLUMNS[col_idx - 1]["key"]
            cell.font = FONT_BODY
            cell.fill = row_fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = _cell_border()

            # Гиперссылки
            if col_key in link_keys and cell.value:
                try:
                    cell.hyperlink = str(cell.value)
                    cell.font = FONT_LINK
                except Exception:
                    pass
            elif col_key in mailto_keys and cell.value:
                try:
                    cell.hyperlink = f"mailto:{cell.value}"
                    cell.font = FONT_LINK
                except Exception:
                    pass

        ws.row_dimensions[row_idx].height = 20

    last_row = len(contacts) + 1  # строки данных + строка заголовка
    _apply_autofilter_freeze(ws, len(COLUMNS), last_row)


def _build_stats_sheet(
    ws,
    categorized: dict[str, list[ContactRecord]],
    task: Optional[ParseTask],
    results: Optional[list[dict[str, Any]]] = None,
) -> None:
    """Заполняет лист «Статистика»."""
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 30

    fill_header = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    fill_label  = PatternFill("solid", fgColor=COLOR_STATS_ROW)
    fill_value  = PatternFill("solid", fgColor=COLOR_SUMMARY_BG)

    def add_section_header(title: str) -> None:
        ws.append([title])
        row = ws.max_row
        cell = ws.cell(row, 1)
        cell.font = FONT_HEADER
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(f"A{row}:B{row}")
        ws.row_dimensions[row].height = 24

    def add_row(label: str, value: Any) -> None:
        ws.append([label, value])
        row = ws.max_row
        ws.cell(row, 1).font = Font(name="Calibri", size=11, bold=True)
        ws.cell(row, 2).font = Font(name="Calibri", size=11)
        ws.cell(row, 1).fill = fill_label
        ws.cell(row, 2).fill = fill_value
        ws.cell(row, 1).border = _cell_border()
        ws.cell(row, 2).border = _cell_border()

    # ── Итоги задачи ──────────────────────────────────────────────────────────
    add_section_header("Итоги парсинга")

    total_contacts = sum(len(v) for v in categorized.values())

    if task:
        add_row("Дата создания задачи", task.created_at.strftime("%d.%m.%Y %H:%M:%S"))
        finished = task.finished_at.strftime("%d.%m.%Y %H:%M:%S") if task.finished_at else "—"
        add_row("Дата завершения", finished)
        add_row("Режим парсинга", f"Режим {task.mode.value}")
        add_row("Вариант извлечения", task.variant.value)
        add_row("Ошибок", task.progress.errors)
        add_row("Просмотрено страниц", task.progress.total_pages)

        if task.progress.elapsed_seconds > 0:
            minutes = int(task.progress.elapsed_seconds // 60)
            seconds = int(task.progress.elapsed_seconds % 60)
            add_row("Время выполнения", f"{minutes} мин {seconds} сек")

        if task.variant.value == "B":
            add_row("Токенов LLM использовано", task.progress.llm_tokens_used)
            add_row("Переключений на Вариант A", task.progress.fallback_count)

    add_row("Итого контактов", total_contacts)

    # ── По категориям ──────────────────────────────────────────────────────────
    ws.append([])
    add_section_header("Контакты по категориям")
    ws.append(["Категория", "Количество"])
    hdr_row = ws.max_row
    ws.cell(hdr_row, 1).font = FONT_HEADER
    ws.cell(hdr_row, 2).font = FONT_HEADER
    ws.cell(hdr_row, 1).fill = PatternFill("solid", fgColor="3D7AB5")
    ws.cell(hdr_row, 2).fill = PatternFill("solid", fgColor="3D7AB5")
    ws.cell(hdr_row, 1).border = _cell_border()
    ws.cell(hdr_row, 2).border = _cell_border()

    for cat, items in sorted(categorized.items(), key=lambda x: -len(x[1])):
        ws.append([cat, len(items)])

    # ── По сайтам (если передан results) ──────────────────────────────────────
    if results:
        ws.append([])
        add_section_header("Результаты по сайтам")
        ws.append(["Сайт", "Контактов найдено"])
        hdr_row = ws.max_row
        ws.cell(hdr_row, 1).font = FONT_HEADER
        ws.cell(hdr_row, 2).font = FONT_HEADER
        ws.cell(hdr_row, 1).fill = PatternFill("solid", fgColor="3D7AB5")
        ws.cell(hdr_row, 2).fill = PatternFill("solid", fgColor="3D7AB5")

        for site_result in sorted(results, key=lambda r: -len(r.get("contacts", []))):
            ws.append([
                site_result.get("site_url", ""),
                len(site_result.get("contacts", [])),
            ])


def _build_log_sheet(
    ws,
    results: list[dict[str, Any]],
    task: Optional[ParseTask],
) -> None:
    """Заполняет лист «Лог» обработки по сайтам."""
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 60

    ws.append(["Время", "Сайт", "Страниц", "Статус / Примечание"])
    header_fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    for cell in ws[1]:
        cell.font = FONT_HEADER
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = _cell_border()

    now_str = datetime.utcnow().strftime("%d.%m.%Y %H:%M")
    for site_result in results:
        contacts_count = len(site_result.get("contacts", []))
        status = "OK" if contacts_count > 0 else "Нет контактов"
        ws.append([
            now_str,
            site_result.get("site_url", ""),
            site_result.get("pages_crawled", 0),
            f"{status}: найдено {contacts_count} контактов",
        ])


# ── Основной класс экспортёра ─────────────────────────────────────────────────

class ExcelExporter:
    """
    Экспортирует результаты парсинга в многолистовой XLSX-файл.

    Каждая CRM-категория становится отдельным листом.
    Дополнительно создаются листы «Статистика» и «Лог».
    """

    async def export(
        self,
        results: list[dict[str, Any]],
        task: ParseTask,
        output_dir: Path,
    ) -> Path:
        """
        Создаёт XLSX-файл с результатами парсинга.

        Args:
            results: Список словарей {site_url, company_name, contacts: [ContactRecord]}.
            task:    Задача парсинга (для метаданных).
            output_dir: Директория для сохранения файла.

        Returns:
            Path к созданному файлу.
        """
        # Собираем все контакты из всех сайтов
        all_contacts: list[ContactRecord] = []
        for site_result in results:
            all_contacts.extend(site_result.get("contacts", []))

        # Категоризируем
        categorized = categorize_contacts(all_contacts)

        wb = openpyxl.Workbook()

        # Удаляем дефолтный лист
        default_sheet = wb.active
        wb.remove(default_sheet)

        # ── Листы по категориям ────────────────────────────────────────────────
        # Порядок: сначала непустые категории из CRM_CATEGORIES, затем «Другие категории»
        from app.core.categorizer import CRM_CATEGORIES, DEFAULT_CATEGORY

        ordered_categories: list[str] = []
        for cat in CRM_CATEGORIES:
            if cat in categorized and categorized[cat]:
                ordered_categories.append(cat)

        # «Другие категории» — всегда последними среди категорийных листов
        if categorized.get(DEFAULT_CATEGORY):
            ordered_categories.append(DEFAULT_CATEGORY)

        for category in ordered_categories:
            contacts = categorized[category]
            if not contacts:
                continue

            sheet_name = _safe_sheet_name(category)
            ws = wb.create_sheet(title=sheet_name)
            _build_category_sheet(ws, contacts)
            logger.debug("Лист «%s»: %d контактов", category, len(contacts))

        # ── Лист «Статистика» ─────────────────────────────────────────────────
        ws_stats = wb.create_sheet("Статистика")
        _build_stats_sheet(ws_stats, categorized, task, results)

        # ── Лист «Лог» ────────────────────────────────────────────────────────
        ws_log = wb.create_sheet("Лог")
        _build_log_sheet(ws_log, results, task)

        # ── Сохранение ────────────────────────────────────────────────────────
        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        task_short = str(task.task_id)[:8]
        filename = f"Результаты_парсинга_{timestamp}_{task_short}.xlsx"
        filepath = output_dir / filename

        wb.save(str(filepath))
        logger.info(
            "Excel сохранён: %s (%.1f KB) — %d листов, %d контактов",
            filepath,
            filepath.stat().st_size / 1024,
            len(wb.sheetnames),
            len(all_contacts),
        )
        return filepath


# ── Утилиты ───────────────────────────────────────────────────────────────────

def _safe_sheet_name(name: str) -> str:
    """
    Приводит имя листа к допустимому для Excel (макс. 31 символ,
    без спецсимволов: / \\ ? * [ ] : ).
    """
    forbidden = re.compile(r"[/\\?*\[\]:]")
    safe = forbidden.sub("_", name)
    return safe[:31]
