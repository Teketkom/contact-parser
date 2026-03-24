"""
REST API и WebSocket эндпоинты приложения Contact Parser.
Обрабатывает создание задач, получение статуса, скачивание результатов,
управление чёрным списком и WebSocket-обновления прогресса в реальном времени.
"""

from __future__ import annotations

import csv
import io
import json
import logging
import os
import uuid
from pathlib import Path
from typing import Optional

import aiofiles
from fastapi import (
    APIRouter,
    BackgroundTasks,
    File,
    Form,
    HTTPException,
    Query,
    UploadFile,
    WebSocket,
    WebSocketDisconnect,
)
from fastapi.responses import FileResponse, JSONResponse

from app.config import settings
from app.core.blacklist import blacklist_manager
from app.core.task_manager import task_manager
from app.models import (
    BlacklistUploadResponse,
    ErrorResponse,
    ExtractionVariant,
    ParseMode,
    ParseTask,
    ParseTaskRequest,
    SiteEntry,
    TaskResponse,
    TaskStatus,
    WSMessage,
    WSMessageType,
)

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api")


# ── Вспомогательные функции ─────────────────────────────────────────────────────────

def _task_to_response(task: ParseTask) -> TaskResponse:
    """Конвертирует ParseTask в TaskResponse для API."""
    return TaskResponse(
        task_id=str(task.task_id),
        status=task.status,
        mode=task.mode,
        variant=task.variant,
        created_at=task.created_at,
        started_at=task.started_at,
        finished_at=task.finished_at,
        progress=task.progress,
        result_file=task.result_file,
        log_file=task.log_file,
        error_message=task.error_message,
    )


def _find_header_row(rows: list) -> int:
    """
    Ищет строку с заголовками в Excel-файле.
    Заголовочная строка содержит хотя бы 2 непустых ячейки и
    хотя бы одно ключевое слово (url, сайт, site, компани, название и т.д.)
    """
    keywords = {"url", "сайт", "site", "компани", "company", "название", "отрасль", "домен", "domain", "адрес", "официальн"}
    for idx, row in enumerate(rows[:10]):  # Проверяем первые 10 строк
        if not row:
            continue
        cells = [str(c).strip().lower() for c in row if c is not None and str(c).strip()]
        if len(cells) < 2:
            continue
        for cell in cells:
            for kw in keywords:
                if kw in cell:
                    return idx
    return 0  # Если заголовки не найдены — считаем что первая строка


def _find_column(headers: list[str], keywords: list[str]) -> Optional[int]:
    """Ищет столбец по списку ключевых слов в заголовках."""
    for i, h in enumerate(headers):
        h_lower = h.lower()
        for kw in keywords:
            if kw in h_lower:
                return i
    return None


def _parse_sites_from_excel_or_csv(content: bytes, filename: str) -> list[SiteEntry]:
    """
    Парсит список сайтов из Excel или CSV файла.
    Автоматически определяет строку заголовков и столбцы с URL, названием компании, ИНН.
    Поддерживает заголовки на русском и английском языках.
    """
    entries: list[SiteEntry] = []

    if filename.endswith((".xlsx", ".xls")):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return entries

            # Автоматически определяем строку заголовков
            header_idx = _find_header_row(rows)
            headers = [str(h).strip().lower() if h else "" for h in rows[header_idx]]

            logger.info(
                "Excel: строка заголовков=%d, заголовки=%s",
                header_idx, headers,
            )

            # Ищем столбец с URL/сайтом — включая "официальный сайт"
            url_col = _find_column(
                headers,
                ["url", "официальн", "сайт", "site", "домен", "domain", "адрес", "web", "ссылка", "link"],
            )

            # Ищем столбец с названием компании
            name_col = _find_column(
                headers,
                ["компани", "company", "название", "организаци", "name", "наименован", "фирма"],
            )

            # Ищем столбец с ИНН
            inn_col = _find_column(headers, ["инн", "inn", "tax"])

            # Если URL-столбец не найден, пробуем найти столбец содержащий URL-подобные значения
            if url_col is None:
                data_rows = rows[header_idx + 1:]
                for col_idx in range(len(headers)):
                    url_like_count = 0
                    for row in data_rows[:5]:
                        if row and col_idx < len(row) and row[col_idx]:
                            val = str(row[col_idx]).strip()
                            if "." in val and " " not in val:
                                url_like_count += 1
                    if url_like_count >= 2:
                        url_col = col_idx
                        break

            if url_col is None:
                url_col = 0
                logger.warning("Не удалось определить столбец URL, используем столбец 0")

            logger.info(
                "Столбцы: url=%s, name=%s, inn=%s",
                url_col, name_col, inn_col,
            )

            for row in rows[header_idx + 1:]:
                if not row or url_col >= len(row) or not row[url_col]:
                    continue
                url_val = str(row[url_col]).strip()
                if not url_val or url_val.startswith("#"):
                    continue
                # Извлечение URL из ячеек с разделителем ; (формат: название;тип;url)
                if ";" in url_val:
                    _parts = [_p.strip() for _p in url_val.split(";")]
                    _url_cand = None
                    for _p in _parts:
                        if _p.startswith(("http://", "https://")) or ("." in _p and " " not in _p and len(_p) > 4):
                            _url_cand = _p
                            break
                    if _url_cand:
                        url_val = _url_cand
                # Добавляем схему если отсутствует
                if not url_val.startswith(("http://", "https://")):
                    url_val = "https://" + url_val
                entry = SiteEntry(
                    url=url_val,
                    company_name=str(row[name_col]).strip() if name_col is not None and name_col < len(row) and row[name_col] else None,
                    inn=str(row[inn_col]).strip() if inn_col is not None and inn_col < len(row) and row[inn_col] else None,
                )
                entries.append(entry)
        except Exception as exc:
            logger.error("Ошибка чтения Excel-файла: %s", exc)
            raise HTTPException(status_code=422, detail=f"Ошибка чтения Excel: {exc}")

    elif filename.endswith(".csv"):
        try:
            text = content.decode("utf-8-sig", errors="replace")
            dialect = csv.Sniffer().sniff(text[:4096], delimiters=",;\t")
            reader = csv.DictReader(io.StringIO(text), dialect=dialect)
            for row in reader:
                url_val = None
                for key in row:
                    if key and any(kw in key.lower() for kw in ["url", "официальн", "сайт", "site", "домен", "domain", "адрес", "web", "ссылка"]):
                        url_val = row[key].strip()
                        break
                if not url_val:
                    for v in row.values():
                        if v and v.strip() and "." in v.strip():
                            url_val = v.strip()
                            break
                if not url_val:
                    url_val = next((v.strip() for v in row.values() if v and v.strip()), None)
                if not url_val:
                    continue
                if not url_val.startswith(("http://", "https://")):
                    url_val = "https://" + url_val

                company = None
                inn = None
                for key in row:
                    if key and any(kw in key.lower() for kw in ["компани", "company", "название", "организаци", "наименован"]):
                        company = row[key].strip() or None
                    if key and ("инн" in key.lower() or "inn" in key.lower()):
                        inn = row[key].strip() or None

                entries.append(SiteEntry(url=url_val, company_name=company, inn=inn))
        except Exception as exc:
            logger.error("Ошибка чтения CSV-файла: %s", exc)
            raise HTTPException(status_code=422, detail=f"Ошибка чтения CSV: {exc}")

    elif filename.endswith(".txt"):
        try:
            text = content.decode("utf-8-sig", errors="replace")
            for line in text.splitlines():
                line = line.strip()
                if not line or line.startswith("#"):
                    continue
                if not line.startswith(("http://", "https://")):
                    line = "https://" + line
                entries.append(SiteEntry(url=line))
        except Exception as exc:
            raise HTTPException(status_code=422, detail=f"Ошибка чтения TXT: {exc}")
    else:
        raise HTTPException(
            status_code=422,
            detail="Поддерживаемые форматы файлов: .xlsx, .csv, .txt",
        )

    return entries


# ── Превью файла ────────────────────────────────────────────────────────────────────

@router.post(
    "/upload/preview",
    summary="Превью загруженного файла",
    tags=["Загрузка"],
)
async def preview_upload(
    file: UploadFile = File(..., description="Excel/CSV/TXT файл для превью"),
) -> dict:
    """
    Валидирует загруженный файл и возвращает превью: количество строк и первые URL.
    """
    content = await file.read()
    filename = file.filename or "upload.csv"

    try:
        sites = _parse_sites_from_excel_or_csv(content, filename)
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=422, detail=f"Ошибка обработки файла: {exc}")

    preview_urls = [s.url for s in sites[:10]]

    return {
        "filename": filename,
        "size_bytes": len(content),
        "rows_count": len(sites),
        "preview_urls": preview_urls,
    }


# ── Управление задачами ────────────────────────────────────────────────────────────

@router.post(
    "/tasks",
    response_model=TaskResponse,
    status_code=201,
    summary="Создать задачу парсинга",
    tags=["Задачи"],
)
async def create_task(
    background_tasks: BackgroundTasks,
    mode: int = Form(..., description="Режим парсинга: 1, 2 или 3"),
    variant: str = Form(default="B", description="Вариант извлечения: A (классический) или B (AI)"),
    target_positions: str = Form(
        default="",
        description="Целевые должности через запятую (для режима 1)",
    ),
    search_queries: str = Form(
        default="",
        description="Поисковые запросы через запятую (для режима 3)",
    ),
    file: Optional[UploadFile] = File(
        default=None,
        description="Excel/CSV/TXT файл со списком сайтов",
    ),
) -> TaskResponse:
    """
    Создаёт новую задачу парсинга контактной информации.
    """
    try:
        parse_mode = ParseMode(mode)
    except ValueError:
        raise HTTPException(status_code=422, detail=f"Неверный режим парсинга: {mode}")

    # Принудительно Вариант B (AI / Perplexity) — игнорируем параметр variant
    extract_variant = ExtractionVariant.AI

    positions = [p.strip() for p in target_positions.split(",") if p.strip()]
    if parse_mode == ParseMode.SITES_WITH_TARGET_POSITIONS and not positions:
        raise HTTPException(
            status_code=422,
            detail="Для режима 1 необходимо указать целевые должности",
        )

    queries = [q.strip() for q in search_queries.split(",") if q.strip()]
    if parse_mode == ParseMode.AUTO_SEARCH and not queries:
        raise HTTPException(
            status_code=422,
            detail="Для режима 3 необходимо указать поисковые запросы",
        )

    sites: list[SiteEntry] = []
    if file is not None:
        content = await file.read()
        sites = _parse_sites_from_excel_or_csv(content, file.filename or "upload.csv")
        if not sites and parse_mode != ParseMode.AUTO_SEARCH:
            raise HTTPException(status_code=422, detail="Файл не содержит URL сайтов")

    if parse_mode in (ParseMode.SITES_WITH_TARGET_POSITIONS, ParseMode.SITES_ALL_POSITIONS) and not sites:
        raise HTTPException(
            status_code=422,
            detail="Для режимов 1 и 2 необходим файл со списком сайтов",
        )

    task = await task_manager.create_task(
        mode=parse_mode,
        variant=extract_variant,
        sites=sites,
        target_positions=positions,
        search_queries=queries,
    )

    background_tasks.add_task(task_manager.run_task, str(task.task_id))

    logger.info(
        "Создана задача %s: режим=%s, вариант=%s, сайтов=%d",
        task.task_id,
        parse_mode,
        extract_variant,
        len(sites),
    )

    return _task_to_response(task)


@router.get(
    "/tasks",
    response_model=list[TaskResponse],
    summary="Список всех задач",
    tags=["Задачи"],
)
async def list_tasks(
    status: Optional[str] = Query(None, description="Фильтр по статусу"),
    limit: int = Query(50, ge=1, le=500),
    offset: int = Query(0, ge=0),
) -> list[TaskResponse]:
    """Возвращает список задач парсинга с фильтрацией по статусу."""
    all_tasks = list(task_manager.tasks.values())

    if status:
        try:
            status_filter = TaskStatus(status)
            all_tasks = [t for t in all_tasks if t.status == status_filter]
        except ValueError:
            pass

    all_tasks.sort(key=lambda t: t.created_at, reverse=True)
    paginated = all_tasks[offset : offset + limit]

    return [_task_to_response(t) for t in paginated]


@router.get(
    "/tasks/{task_id}",
    response_model=TaskResponse,
    summary="Статус задачи",
    tags=["Задачи"],
)
async def get_task(task_id: str) -> TaskResponse:
    """Возвращает текущий статус и прогресс задачи парсинга."""
    task = task_manager.tasks.get(task_id)
    if not task:
        raise HTTPException(status_code=404, detail=f"Задача {task_id} не найдена")
    return _task_to_response(task)


@router.delete(
    "/tasks/{task_id}",
    summary="Отменить/удалить задачу",
    tags=["Задачи"],
)
async def cancel_or_delete_task(task_id: str) -> dict:
    """Отменяет выполнение или удаляет задачу парсинга."""
    task = task_manager.tasks.get(task_id)
    if not task:
        raise HTTPException(status_code=404, detail=f"Задача {task_id} не найдена")
    if task.status in (TaskStatus.PENDING, TaskStatus.RUNNING):
        await task_manager.cancel_task(task_id)
        return {"message": f"Задача {task_id} отменена"}
    else:
        del task_manager.tasks[task_id]
        return {"message": f"Задача {task_id} удалена"}


# ── Результаты и логи ──────────────────────────────────────────────────────────────

@router.get(
    "/tasks/{task_id}/results",
    summary="Скачать результаты в Excel",
    tags=["Результаты"],
)
async def download_results(task_id: str) -> FileResponse:
    """Скачивает результаты парсинга в формате Excel (.xlsx)."""
    task = task_manager.tasks.get(task_id)
    if not task:
        raise HTTPException(status_code=404, detail=f"Задача {task_id} не найдена")

    if task.status != TaskStatus.COMPLETED:
        raise HTTPException(
            status_code=400,
            detail=f"Задача ещё не завершена (статус: {task.status})",
        )

    if not task.result_file:
        raise HTTPException(status_code=404, detail="Файл результатов не найден")

    result_path = Path(task.result_file)
    if not result_path.exists():
        raise HTTPException(status_code=404, detail="Файл результатов отсутствует на диске")

    return FileResponse(
        path=str(result_path),
        filename=result_path.name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@router.get(
    "/tasks/{task_id}/download",
    summary="Скачать результаты (алиас)",
    tags=["Результаты"],
    include_in_schema=False,
)
async def download_results_alias(task_id: str) -> FileResponse:
    """Алиас для обратной совместимости."""
    return await download_results(task_id)


@router.get(
    "/tasks/{task_id}/logs",
    summary="Скачать лог ошибок",
    tags=["Результаты"],
)
async def download_logs(task_id: str) -> FileResponse:
    """Скачивает файл логов ошибок для задачи парсинга."""
    task = task_manager.tasks.get(task_id)
    if not task:
        raise HTTPException(status_code=404, detail=f"Задача {task_id} не найдена")

    if not task.log_file:
        raise HTTPException(status_code=404, detail="Лог-файл не найден")

    log_path = Path(task.log_file)
    if not log_path.exists():
        raise HTTPException(status_code=404, detail="Лог-файл отсутствует на диске")

    return FileResponse(
        path=str(log_path),
        filename=log_path.name,
        media_type="text/plain; charset=utf-8",
    )



# ── Предпросмотр записей ──────────────────────────────────────────────────────────

@router.get(
    "/tasks/{task_id}/records",
    summary="Предпросмотр записей контактов",
    tags=["Результаты"],
)
async def get_task_records(
    task_id: str,
    limit: int = Query(20, ge=1, le=200, description="Максимум записей"),
    offset: int = Query(0, ge=0, description="Смещение"),
) -> list[dict]:
    """
    Возвращает записи контактов из результатов парсинга для предпросмотра в UI.
    Читает данные из сохранённого Excel-файла.
    """
    task = task_manager.tasks.get(task_id)
    if not task:
        raise HTTPException(status_code=404, detail=f"Задача {task_id} не найдена")

    if task.status != TaskStatus.COMPLETED or not task.result_file:
        return []

    result_path = Path(task.result_file)
    if not result_path.exists():
        return []

    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(result_path), read_only=True)
        ws = wb["Контакты"]

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            wb.close()
            return []

        headers = [str(h).strip() if h else "" for h in rows[0]]
        records = []

        header_map = {
            "Название компании": "company_name",
            "Сайт": "site_url",
            "Общий email": "company_email",
            "Должность (как на сайте)": "position_raw",
            "Должность (нормализованная)": "position_normalized",
            "Должность (норм.)": "position_normalized",
            "ФИО": "full_name",
            "Личный email": "personal_email",
            "Телефон": "phone",
            "ИНН": "inn",
            "КПП": "kpp",
            "ВКонтакте": "social_vk",
            "Telegram": "social_telegram",
            "LinkedIn": "social_linkedin",
            "Соцсети (прочие)": "social_other",
            "URL страницы-источника": "source_url",
            "URL-источник": "source_url",
            "Язык страницы": "page_language",
            "Дата сканирования": "scan_date",
            "Вариант извлечения": "extraction_variant",
            "Вариант": "extraction_variant",
            "Статус обработки": "status",
            "Статус": "status",
            "Комментарий": "comment",
        }

        col_mapping: list[tuple[int, str]] = []
        for i, h in enumerate(headers):
            key = header_map.get(h)
            if key:
                col_mapping.append((i, key))

        data_rows = rows[1 + offset : 1 + offset + limit]

        for row in data_rows:
            record = {}
            for col_idx, key in col_mapping:
                val = row[col_idx] if col_idx < len(row) else None
                record[key] = str(val) if val is not None else None
            records.append(record)

        wb.close()
        return records

    except Exception as exc:
        logger.warning("Ошибка чтения записей из %s: %s", result_path, exc)
        return []


# ── Чёрный список ──────────────────────────────────────────────────────────────────

@router.post(
    "/blacklist",
    response_model=BlacklistUploadResponse,
    summary="Загрузить чёрный список",
    tags=["Чёрный список"],
)
async def upload_blacklist(
    file: UploadFile = File(..., description="Файл чёрного списка (txt/csv/xlsx)"),
) -> BlacklistUploadResponse:
    """
    Загружает список доменов/email/ИНН, которые должны быть исключены из парсинга.
    """
    content = await file.read()
    filename = file.filename or "blacklist.txt"
    added = await blacklist_manager.load_from_file(content, filename)
    total = blacklist_manager.total_count()

    return BlacklistUploadResponse(
        added=added,
        total=total,
        message=f"Добавлено {added} новых записей. Всего в списке: {total}",
    )


@router.get(
    "/blacklist",
    summary="Статистика чёрного списка",
    tags=["Чёрный список"],
)
async def get_blacklist_stats() -> dict:
    """Возвращает статистику чёрного списка."""
    return {
        "domains": blacklist_manager.count("domains"),
        "emails": blacklist_manager.count("emails"),
        "inns": blacklist_manager.count("inns"),
        "total": blacklist_manager.total_count(),
    }


@router.delete(
    "/blacklist",
    summary="Очистить чёрный список",
    tags=["Чёрный список"],
)
async def clear_blacklist() -> dict:
    """Полностью очищает чёрный список."""
    blacklist_manager.clear()
    return {"message": "Чёрный список очищен"}


# ── WebSocket ────────────────────────────────────────────────────────────────────────

@router.websocket("/ws/{task_id}")
async def websocket_task_progress(websocket: WebSocket, task_id: str) -> None:
    """
    WebSocket-соединение для получения обновлений прогресса задачи в реальном времени.
    """
    await websocket.accept()

    task = task_manager.tasks.get(task_id)
    if not task:
        await websocket.send_json(
            WSMessage(
                type=WSMessageType.ERROR,
                task_id=task_id,
                data={"message": f"Задача {task_id} не найдена"},
            ).model_dump(mode="json")
        )
        await websocket.close(code=4004)
        return

    logger.info("WebSocket подключён к задаче %s", task_id)

    # Отправляем текущее состояние сразу при подключении
    await websocket.send_json(
        WSMessage(
            type=WSMessageType.PROGRESS,
            task_id=task_id,
            data=task.progress.model_dump(),
        ).model_dump(mode="json")
    )

    # Регистрируем подписчика
    queue = await task_manager.subscribe(task_id)

    try:
        while True:
            try:
                message = await queue.get()
                await websocket.send_json(message)

                if message.get("type") in (
                    WSMessageType.COMPLETED,
                    WSMessageType.ERROR,
                    WSMessageType.CANCELLED,
                ):
                    break
            except Exception as exc:
                logger.warning("Ошибка отправки WebSocket-сообщения: %s", exc)
                break
    except WebSocketDisconnect:
        logger.info("WebSocket отключён от задачи %s", task_id)
    finally:
        await task_manager.unsubscribe(task_id, queue)
        try:
            await websocket.close()
        except Exception:
            pass
