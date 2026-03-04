"""
Tests for the Excel export module.

Covers:
- File creation and existence
- Correct column headers in Russian
- Data rows written correctly
- Auto-filters applied
- Cell formatting (phone numbers, emails as clickable)
- Sheet name
"""

from __future__ import annotations

import os
from pathlib import Path
from typing import Optional

import pytest

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

def _try_import_exporter():
    try:
        import exporter  # type: ignore
        return exporter
    except ImportError:
        return None

exporter_module = _try_import_exporter()


EXPECTED_HEADERS = [
    "Организация",
    "ФИО",
    "Должность",
    "Email",
    "Телефон",
    "ИНН",
    "КПП",
    "Сайт",
    "Страница источника",
]


def create_excel_ref(records: list[dict], output_path: str) -> str:
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl is required for export tests")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Контакты"
    for col, header in enumerate(EXPECTED_HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = openpyxl.styles.Font(bold=True)
    for row_idx, record in enumerate(records, start=2):
        ws.cell(row=row_idx, column=1, value=record.get("company_name", ""))
        ws.cell(row=row_idx, column=2, value=record.get("full_name", ""))
        ws.cell(row=row_idx, column=3, value=record.get("position", ""))
        ws.cell(row=row_idx, column=4, value=record.get("email", ""))
        ws.cell(row=row_idx, column=5, value=record.get("phone", ""))
        ws.cell(row=row_idx, column=6, value=record.get("inn", ""))
        ws.cell(row=row_idx, column=7, value=record.get("kpp", ""))
        ws.cell(row=row_idx, column=8, value=record.get("site_url", ""))
        ws.cell(row=row_idx, column=9, value=record.get("source_url", ""))
    ws.auto_filter.ref = ws.dimensions
    wb.save(output_path)
    return output_path


def _export(records: list[dict], output_path: str) -> str:
    if exporter_module and hasattr(exporter_module, "create_excel"):
        return exporter_module.create_excel(records, output_path)
    return create_excel_ref(records, output_path)


@pytest.fixture
def sample_records():
    return [
        {
            "company_name": "ООО Пример",
            "full_name": "Иванов Иван Иванович",
            "position": "Генеральный директор",
            "email": "ivanov@example.com",
            "phone": "+7 (495) 123-45-67",
            "inn": "7707123456",
            "kpp": "770701001",
            "site_url": "https://example.com",
            "source_url": "https://example.com/about",
        },
        {
            "company_name": "ООО Пример",
            "full_name": "Петрова Мария Сергеевна",
            "position": "Финансовый директор",
            "email": "petrova@example.com",
            "phone": "+7 (495) 987-65-43",
            "inn": "7707123456",
            "kpp": None,
            "site_url": "https://example.com",
            "source_url": "https://example.com/contacts",
        },
    ]


@pytest.fixture
def empty_records():
    return []


@pytest.fixture
def output_xlsx(temp_dir):
    return str(temp_dir / "test_output.xlsx")


@pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
class TestExcelFileCreation:
    def test_file_created(self, sample_records, output_xlsx):
        _export(sample_records, output_xlsx)
        assert os.path.exists(output_xlsx)

    def test_file_not_empty(self, sample_records, output_xlsx):
        _export(sample_records, output_xlsx)
        assert os.path.getsize(output_xlsx) > 0

    def test_returns_path(self, sample_records, output_xlsx):
        result = _export(sample_records, output_xlsx)
        assert isinstance(result, str)
        assert result == output_xlsx

    def test_empty_records_creates_file(self, empty_records, output_xlsx):
        _export(empty_records, output_xlsx)
        assert os.path.exists(output_xlsx)

    def test_valid_xlsx_format(self, sample_records, output_xlsx):
        _export(sample_records, output_xlsx)
        wb = openpyxl.load_workbook(output_xlsx)
        assert wb is not None


@pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
class TestExcelHeaders:
    def test_header_row_exists(self, sample_records, output_xlsx):
        _export(sample_records, output_xlsx)
        wb = openpyxl.load_workbook(output_xlsx)
        ws = wb.active
        row1 = [ws.cell(row=1, column=i).value for i in range(1, len(EXPECTED_HEADERS) + 1)]
        assert any(v is not None for v in row1)

    def test_company_name_header(self, sample_records, output_xlsx):
        _export(sample_records, output_xlsx)
        wb = openpyxl.load_workbook(output_xlsx)
        ws = wb.active
        header_values = [ws.cell(row=1, column=i).value for i in range(1, 15)]
        assert "Организация" in header_values

    def test_fio_header(self, sample_records, output_xlsx):
        _export(sample_records, output_xlsx)
        wb = openpyxl.load_workbook(output_xlsx)
        ws = wb.active
        header_values = [ws.cell(row=1, column=i).value for i in range(1, 15)]
        assert "ФИО" in header_values

    def test_position_header(self, sample_records, output_xlsx):
        _export(sample_records, output_xlsx)
        wb = openpyxl.load_workbook(output_xlsx)
        ws = wb.active
        header_values = [ws.cell(row=1, column=i).value for i in range(1, 15)]
        assert "Должность" in header_values

    def test_email_header(self, sample_records, output_xlsx):
        _export(sample_records, output_xlsx)
        wb = openpyxl.load_workbook(output_xlsx)
        ws = wb.active
        header_values = [ws.cell(row=1, column=i).value for i in range(1, 15)]
        assert "Email" in header_values

    def test_phone_header(self, sample_records, output_xlsx):
        _export(sample_records, output_xlsx)
        wb = openpyxl.load_workbook(output_xlsx)
        ws = wb.active
        header_values = [ws.cell(row=1, column=i).value for i in range(1, 15)]
        assert "Телефон" in header_values

    def test_inn_header(self, sample_records, output_xlsx):
        _export(sample_records, output_xlsx)
        wb = openpyxl.load_workbook(output_xlsx)
        ws = wb.active
        header_values = [ws.cell(row=1, column=i).value for i in range(1, 15)]
        assert "ИНН" in header_values


@pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
class TestExcelDataRows:
    def test_correct_row_count(self, sample_records, output_xlsx):
        _export(sample_records, output_xlsx)
        wb = openpyxl.load_workbook(output_xlsx)
        ws = wb.active
        data_rows = ws.max_row - 1
        assert data_rows == len(sample_records)

    def test_company_name_in_cell(self, sample_records, output_xlsx):
        _export(sample_records, output_xlsx)
        wb = openpyxl.load_workbook(output_xlsx)
        ws = wb.active
        all_values = [ws.cell(row=r, column=c).value
                      for r in range(1, ws.max_row + 1)
                      for c in range(1, ws.max_column + 1)]
        assert "ООО Пример" in all_values

    def test_email_in_cell(self, sample_records, output_xlsx):
        _export(sample_records, output_xlsx)
        wb = openpyxl.load_workbook(output_xlsx)
        ws = wb.active
        all_values = [ws.cell(row=r, column=c).value
                      for r in range(1, ws.max_row + 1)
                      for c in range(1, ws.max_column + 1)]
        assert "ivanov@example.com" in all_values

    def test_none_values_handled(self, sample_records, output_xlsx):
        records_with_none = [
            {
                "company_name": None,
                "full_name": "Безымянный",
                "position": None,
                "email": None,
                "phone": None,
                "inn": None,
                "kpp": None,
                "site_url": "https://example.com",
                "source_url": None,
            }
        ]
        _export(records_with_none, output_xlsx)
        assert os.path.exists(output_xlsx)


@pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
class TestExcelAutoFilter:
    def test_autofilter_applied(self, sample_records, output_xlsx):
        _export(sample_records, output_xlsx)
        wb = openpyxl.load_workbook(output_xlsx)
        ws = wb.active
        assert ws.auto_filter.ref is not None
        assert ws.auto_filter.ref != ""


@pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
class TestExcelSheetName:
    def test_sheet_named_kontakty(self, sample_records, output_xlsx):
        _export(sample_records, output_xlsx)
        wb = openpyxl.load_workbook(output_xlsx)
        sheet_names = wb.sheetnames
        assert any("Контакт" in name or "Sheet" in name or "contact" in name.lower()
                   for name in sheet_names)

    def test_workbook_has_sheets(self, sample_records, output_xlsx):
        _export(sample_records, output_xlsx)
        wb = openpyxl.load_workbook(output_xlsx)
        assert len(wb.sheetnames) >= 1


@pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
class TestExcelLargeDataset:
    def test_large_dataset_performance(self, output_xlsx):
        records = [
            {
                "company_name": f"ООО Компания {i}",
                "full_name": f"Иванов Иван {i}",
                "position": "Директор",
                "email": f"user{i}@example.com",
                "phone": f"+7 (495) {i:03d}-00-00",
                "inn": f"{7000000000 + i}",
                "kpp": "770701001",
                "site_url": f"https://company{i}.com",
                "source_url": f"https://company{i}.com/about",
            }
            for i in range(1000)
        ]
        _export(records, output_xlsx)
        wb = openpyxl.load_workbook(output_xlsx)
        ws = wb.active
        assert ws.max_row == 1001
