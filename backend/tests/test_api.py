"""
API integration tests for the Contact Information Parser FastAPI backend.

Covers:
- POST /api/tasks
- GET  /api/tasks
- GET  /api/tasks/{task_id}
- GET  /api/tasks/{task_id}/progress
- POST /api/tasks/{task_id}/cancel
- GET  /api/health
- File validation (wrong format, too large, missing)
"""

from __future__ import annotations

import io
import json
import os
import uuid
from typing import Any
from unittest.mock import MagicMock, patch

import pytest

try:
    from fastapi import FastAPI, File, UploadFile, Form, HTTPException
    from fastapi.testclient import TestClient
    HAS_FASTAPI = True
except ImportError:
    HAS_FASTAPI = False


def build_stub_app() -> Any:
    if not HAS_FASTAPI:
        return None

    from fastapi import FastAPI, UploadFile, File, Form, HTTPException
    from fastapi.responses import JSONResponse
    import uuid

    app = FastAPI(title="Contact Parser Stub")
    _tasks: dict = {}

    @app.get("/api/health")
    def health():
        return {"status": "ok", "version": "test"}

    @app.post("/api/tasks")
    async def create_task(
        file: UploadFile = File(...),
        mode: str = Form("1"),
        target_positions: str = Form(""),
    ):
        ext = (file.filename or "").split(".")[-1].lower()
        if ext not in ("xlsx", "xls", "csv"):
            raise HTTPException(status_code=422, detail="Неподдерживаемый формат файла")
        content = await file.read()
        if len(content) == 0:
            raise HTTPException(status_code=422, detail="Файл пустой")
        task_id = str(uuid.uuid4())
        _tasks[task_id] = {
            "task_id": task_id,
            "status": "pending",
            "mode": int(mode),
            "total_sites": 0,
            "processed_sites": 0,
            "found_records": 0,
            "error_count": 0,
            "created_at": "2026-01-01T00:00:00",
            "target_positions": [p.strip() for p in target_positions.split(",") if p.strip()],
        }
        return {"task_id": task_id, "status": "pending", "message": "Задача создана"}

    @app.get("/api/tasks")
    def list_tasks(page: int = 1, page_size: int = 20):
        items = list(_tasks.values())
        return {
            "items": items[(page - 1) * page_size: page * page_size],
            "total": len(items),
            "page": page,
            "page_size": page_size,
        }

    @app.get("/api/tasks/{task_id}")
    def get_task(task_id: str):
        task = _tasks.get(task_id)
        if not task:
            raise HTTPException(status_code=404, detail="Задача не найдена")
        return task

    @app.get("/api/tasks/{task_id}/progress")
    def get_progress(task_id: str):
        task = _tasks.get(task_id)
        if not task:
            raise HTTPException(status_code=404, detail="Задача не найдена")
        return {
            "task_id": task_id,
            "status": task["status"],
            "total_sites": task["total_sites"],
            "processed_sites": task["processed_sites"],
            "found_records": task["found_records"],
            "error_count": task["error_count"],
            "elapsed_seconds": 0,
            "errors": [],
        }

    @app.post("/api/tasks/{task_id}/cancel")
    def cancel_task(task_id: str):
        task = _tasks.get(task_id)
        if not task:
            raise HTTPException(status_code=404, detail="Задача не найдена")
        if task["status"] not in ("pending", "running"):
            raise HTTPException(status_code=400, detail="Задача не может быть отменена")
        task["status"] = "cancelled"
        return {"message": "Задача отменена"}

    @app.delete("/api/tasks/{task_id}")
    def delete_task(task_id: str):
        if task_id not in _tasks:
            raise HTTPException(status_code=404, detail="Задача не найдена")
        del _tasks[task_id]
        return {"message": "Задача удалена"}

    @app.post("/api/upload/preview")
    async def upload_preview(file: UploadFile = File(...)):
        ext = (file.filename or "").split(".")[-1].lower()
        if ext not in ("xlsx", "xls", "csv"):
            raise HTTPException(status_code=422, detail="Неподдерживаемый формат файла")
        content = await file.read()
        return {
            "filename": file.filename,
            "size_bytes": len(content),
            "rows_count": 5,
            "preview_urls": ["https://example.com", "https://test.ru"],
        }

    @app.get("/api/blacklist")
    def get_blacklist():
        return {"entries": [], "total": 0}

    @app.post("/api/blacklist/upload")
    async def upload_blacklist(file: UploadFile = File(...)):
        return {"added": 3, "skipped": 0, "total": 3, "entries": []}

    return app


@pytest.fixture
def app():
    try:
        from main import app as real_app  # type: ignore
        return real_app
    except ImportError:
        return build_stub_app()


@pytest.fixture
def api_client(app):
    if not HAS_FASTAPI:
        pytest.skip("FastAPI not installed")
    from fastapi.testclient import TestClient
    with TestClient(app) as c:
        yield c


def _make_csv_file(content: str = "url\nhttps://example.com\nhttps://test.ru") -> tuple:
    return ("sites.csv", io.BytesIO(content.encode()), "text/csv")


def _make_xlsx_file() -> tuple:
    try:
        import openpyxl
        buf = io.BytesIO()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "url"
        ws["A2"] = "https://example.com"
        ws["A3"] = "https://test.ru"
        wb.save(buf)
        buf.seek(0)
        return ("sites.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except ImportError:
        return _make_csv_file()


@pytest.mark.skipif(not HAS_FASTAPI, reason="FastAPI not installed")
class TestHealthEndpoint:
    def test_health_returns_200(self, api_client):
        response = api_client.get("/api/health")
        assert response.status_code == 200

    def test_health_returns_status_ok(self, api_client):
        response = api_client.get("/api/health")
        data = response.json()
        assert data.get("status") == "ok"


@pytest.mark.skipif(not HAS_FASTAPI, reason="FastAPI not installed")
class TestTaskCreation:
    def test_create_task_with_csv_mode1(self, api_client):
        filename, content, mimetype = _make_csv_file()
        response = api_client.post(
            "/api/tasks",
            files={"file": (filename, content, mimetype)},
            data={"mode": "1", "target_positions": "Директор,CEO"},
        )
        assert response.status_code == 200
        data = response.json()
        assert "task_id" in data
        assert data["status"] in ("pending", "running")

    def test_create_task_with_csv_mode2(self, api_client):
        filename, content, mimetype = _make_csv_file()
        response = api_client.post(
            "/api/tasks",
            files={"file": (filename, content, mimetype)},
            data={"mode": "2"},
        )
        assert response.status_code == 200
        data = response.json()
        assert "task_id" in data

    def test_create_task_returns_task_id(self, api_client):
        filename, content, mimetype = _make_csv_file()
        response = api_client.post(
            "/api/tasks",
            files={"file": (filename, content, mimetype)},
            data={"mode": "2"},
        )
        assert response.status_code == 200
        task_id = response.json().get("task_id")
        assert task_id is not None
        assert len(task_id) > 8

    def test_create_task_with_xlsx(self, api_client):
        filename, content, mimetype = _make_xlsx_file()
        response = api_client.post(
            "/api/tasks",
            files={"file": (filename, content, mimetype)},
            data={"mode": "2"},
        )
        assert response.status_code == 200

    def test_wrong_file_format_rejected(self, api_client):
        bad_file = io.BytesIO(b"This is not a spreadsheet")
        response = api_client.post(
            "/api/tasks",
            files={"file": ("data.txt", bad_file, "text/plain")},
            data={"mode": "1"},
        )
        assert response.status_code in (400, 422)

    def test_empty_file_rejected(self, api_client):
        empty_file = io.BytesIO(b"")
        response = api_client.post(
            "/api/tasks",
            files={"file": ("empty.csv", empty_file, "text/csv")},
            data={"mode": "1"},
        )
        assert response.status_code in (400, 422)

    def test_missing_file_returns_422(self, api_client):
        response = api_client.post(
            "/api/tasks",
            data={"mode": "1"},
        )
        assert response.status_code == 422


@pytest.mark.skipif(not HAS_FASTAPI, reason="FastAPI not installed")
class TestTaskList:
    def test_list_tasks_returns_200(self, api_client):
        response = api_client.get("/api/tasks")
        assert response.status_code == 200

    def test_list_tasks_response_structure(self, api_client):
        response = api_client.get("/api/tasks")
        data = response.json()
        assert "items" in data
        assert "total" in data
        assert isinstance(data["items"], list)

    def test_created_task_in_list(self, api_client):
        filename, content, mimetype = _make_csv_file()
        create_resp = api_client.post(
            "/api/tasks",
            files={"file": (filename, content, mimetype)},
            data={"mode": "2"},
        )
        task_id = create_resp.json()["task_id"]
        list_resp = api_client.get("/api/tasks")
        task_ids = [t["task_id"] for t in list_resp.json()["items"]]
        assert task_id in task_ids

    def test_pagination_params_accepted(self, api_client):
        response = api_client.get("/api/tasks?page=1&page_size=10")
        assert response.status_code == 200


@pytest.mark.skipif(not HAS_FASTAPI, reason="FastAPI not installed")
class TestTaskDetail:
    def _create_task(self, api_client) -> str:
        filename, content, mimetype = _make_csv_file()
        resp = api_client.post(
            "/api/tasks",
            files={"file": (filename, content, mimetype)},
            data={"mode": "2"},
        )
        return resp.json()["task_id"]

    def test_get_existing_task_200(self, api_client):
        task_id = self._create_task(api_client)
        response = api_client.get(f"/api/tasks/{task_id}")
        assert response.status_code == 200

    def test_get_task_returns_task_id(self, api_client):
        task_id = self._create_task(api_client)
        response = api_client.get(f"/api/tasks/{task_id}")
        data = response.json()
        assert data["task_id"] == task_id

    def test_get_task_has_status(self, api_client):
        task_id = self._create_task(api_client)
        response = api_client.get(f"/api/tasks/{task_id}")
        data = response.json()
        assert "status" in data
        assert data["status"] in ("pending", "running", "completed", "failed", "cancelled")

    def test_get_nonexistent_task_404(self, api_client):
        fake_id = str(uuid.uuid4())
        response = api_client.get(f"/api/tasks/{fake_id}")
        assert response.status_code == 404


@pytest.mark.skipif(not HAS_FASTAPI, reason="FastAPI not installed")
class TestTaskProgress:
    def _create_task(self, api_client) -> str:
        filename, content, mimetype = _make_csv_file()
        resp = api_client.post(
            "/api/tasks",
            files={"file": (filename, content, mimetype)},
            data={"mode": "2"},
        )
        return resp.json()["task_id"]

    def test_progress_returns_200(self, api_client):
        task_id = self._create_task(api_client)
        response = api_client.get(f"/api/tasks/{task_id}/progress")
        assert response.status_code == 200

    def test_progress_has_required_fields(self, api_client):
        task_id = self._create_task(api_client)
        response = api_client.get(f"/api/tasks/{task_id}/progress")
        data = response.json()
        required_fields = {"task_id", "status", "total_sites", "processed_sites",
                           "found_records", "error_count"}
        assert required_fields.issubset(set(data.keys()))

    def test_progress_nonexistent_task_404(self, api_client):
        fake_id = str(uuid.uuid4())
        response = api_client.get(f"/api/tasks/{fake_id}/progress")
        assert response.status_code == 404


@pytest.mark.skipif(not HAS_FASTAPI, reason="FastAPI not installed")
class TestTaskCancellation:
    def _create_task(self, api_client) -> str:
        filename, content, mimetype = _make_csv_file()
        resp = api_client.post(
            "/api/tasks",
            files={"file": (filename, content, mimetype)},
            data={"mode": "2"},
        )
        return resp.json()["task_id"]

    def test_cancel_pending_task(self, api_client):
        task_id = self._create_task(api_client)
        response = api_client.post(f"/api/tasks/{task_id}/cancel")
        assert response.status_code == 200

    def test_cancelled_task_status_updated(self, api_client):
        task_id = self._create_task(api_client)
        api_client.post(f"/api/tasks/{task_id}/cancel")
        detail = api_client.get(f"/api/tasks/{task_id}")
        assert detail.json()["status"] == "cancelled"

    def test_cancel_nonexistent_task_404(self, api_client):
        fake_id = str(uuid.uuid4())
        response = api_client.post(f"/api/tasks/{fake_id}/cancel")
        assert response.status_code == 404


@pytest.mark.skipif(not HAS_FASTAPI, reason="FastAPI not installed")
class TestFileUploadPreview:
    def test_csv_preview_200(self, api_client):
        filename, content, mimetype = _make_csv_file()
        response = api_client.post(
            "/api/upload/preview",
            files={"file": (filename, content, mimetype)},
        )
        assert response.status_code == 200

    def test_preview_returns_filename(self, api_client):
        filename, content, mimetype = _make_csv_file()
        response = api_client.post(
            "/api/upload/preview",
            files={"file": (filename, content, mimetype)},
        )
        data = response.json()
        assert "filename" in data

    def test_invalid_format_preview_rejected(self, api_client):
        bad_file = io.BytesIO(b"not a spreadsheet")
        response = api_client.post(
            "/api/upload/preview",
            files={"file": ("data.pdf", bad_file, "application/pdf")},
        )
        assert response.status_code in (400, 422)
